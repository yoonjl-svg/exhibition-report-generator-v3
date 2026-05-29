"""
Excel 일괄 임포트 — 템플릿 생성 + 파싱.

큐레이터가 외부에서 데이터를 정리한 뒤 한 번에 업로드할 수 있는
표준 Excel 템플릿을 제공. 폼 입력의 대체가 아니라 보조 경로.

워크플로:
  1. 워크스페이스에서 'Excel 템플릿 다운로드' → 엑셀 파일 받음
  2. 큐레이터가 셀에 데이터 작성 후 저장
  3. 워크스페이스에서 'Excel 가져오기' → 업로드
  4. 파싱 결과로 폼이 미리 채워진 채 상세 모드 진입
  5. 검수·수정 후 'KB에 저장' 클릭

시트 구성:
  1. 기본정보 — 스칼라 (제목·기간·작가·기획진 등)
  2. 정량데이터 — 스칼라 숫자 (예산·관객·작품·프로그램 등)
  3. 프로그램 — 행 단위 리스트
  4. 인쇄물 — 행 단위 리스트
  5. 언론보도 — 행 단위 리스트 (인쇄/온라인 type 컬럼)
  6. 관객후기 — 행 단위 리스트
"""

import io
from datetime import datetime, date as _date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# 시트 1: 기본정보 — 필드명 → session_state 키
# ──────────────────────────────────────────────

BASIC_FIELDS = [
    # (라벨, key, 설명/예시)
    ("전시 제목", "exhibition_title", ""),
    ("전시 유형 (1=정기기획전 / 2=특별전 / 3=기타)", "__exhibition_type", "1"),
    ("전시 시작일 (YYYY-MM-DD)", "period_start", "2025-06-13"),
    ("전시 종료일 (YYYY-MM-DD)", "period_end", "2025-08-17"),
    ("참여 작가 (쉼표 구분)", "artists", ""),
    ("책임기획", "chief_curator", ""),
    ("학예팀", "curatorial_team", ""),
    ("기획", "curators", ""),
    ("홍보", "pr_person", ""),
    ("진행", "coordinators", ""),
    ("후원", "sponsors", ""),
    ("그래픽 디자인", "graphic_designer", ""),
    ("공간 구성", "space_designer", ""),
    ("전시 에세이 (긴 문장 가능)", "theme_text", ""),
    ("멤버십 커뮤니케이션", "membership_text", ""),
    ("광고", "promo_advertising", ""),
    ("보도자료", "promo_press_release", ""),
    ("웹 초청장", "promo_web_invitation", ""),
    ("뉴스레터", "promo_newsletter", ""),
    ("SNS", "promo_sns", ""),
    ("그 외 홍보", "promo_other", ""),
]


# ──────────────────────────────────────────────
# 시트 2: 정량데이터 — 필드명 → session_state 키
# ──────────────────────────────────────────────

QUANTITATIVE_FIELDS = [
    # 예산 및 수입
    ("전시 사용 예산 (원)", "budget_exhibition"),
    ("부대 사용 예산 (원)", "budget_supplementary"),
    ("예산 계획액 (원)", "budget_planned"),
    ("입장 수입 (원)", "ticket_revenue"),
    ("기타 수입 (원)", "other_revenue"),
    # 관객
    ("총 관객수", "total_visitors"),
    ("일반 관객수", "visitor_general"),
    ("학생 관객수", "visitor_student"),
    ("초대 관객수", "visitor_invitation"),
    ("예술인패스 관객수", "visitor_artpass"),
    ("디스커버서울패스 관객수", "visitor_discover"),
    ("기타 할인 관객수", "visitor_discount"),
    ("단체 관객수", "visitor_group"),
    ("오프닝 참석 인원", "opening_attendance"),
    # 작품
    ("출품 작품 - 회화", "artwork_painting"),
    ("출품 작품 - 조각", "artwork_sculpture"),
    ("출품 작품 - 사진", "artwork_photo"),
    ("출품 작품 - 설치", "artwork_installation"),
    ("출품 작품 - 미디어", "artwork_media"),
    ("출품 작품 - 기타", "artwork_other"),
    # 프로그램
    ("프로그램 총 수", "program_count"),
    ("프로그램 총 회차", "program_sessions"),
    ("프로그램 참여 인원", "program_participants"),
    ("도슨트 참여 인원 (총)", "docent_total"),
    ("정기 도슨트", "docent_regular"),
    ("특별 도슨트", "docent_special"),
    # 인력
    ("운영 인력 총원", "staff_total"),
    ("유급 스태프", "staff_paid"),
    ("봉사자", "staff_volunteer"),
    # 홍보 지표
    ("언론 보도 건수", "press_count"),
    ("웹 초청장 발송 수", "web_invitation_count"),
    ("뉴스레터 오픈율 (%)", "newsletter_open_rate"),
    ("SNS 게시 건수", "sns_posts"),
    ("SNS 피드백 합계", "sns_feedback"),
    ("멤버십 회원수", "membership_count"),
]


# ──────────────────────────────────────────────
# 시트 스타일
# ──────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="255C4A")  # 미술관 톤 녹색
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FILL = PatternFill("solid", fgColor="EEF2EA")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DDD4"),
    right=Side(style="thin", color="D9DDD4"),
    top=Side(style="thin", color="D9DDD4"),
    bottom=Side(style="thin", color="D9DDD4"),
)


def _style_header_row(ws, row_num: int, num_cols: int):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_label_cell(cell):
    cell.fill = LABEL_FILL
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def _style_value_cell(cell):
    cell.font = Font(size=10)
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


# ──────────────────────────────────────────────
# 템플릿 생성
# ──────────────────────────────────────────────

def generate_template_xlsx() -> bytes:
    """v5 임포트용 표준 Excel 템플릿을 바이트로 반환."""
    wb = openpyxl.Workbook()

    # ── 시트 1: 기본정보 ──
    ws1 = wb.active
    ws1.title = "기본정보"
    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 60
    ws1["A1"] = "필드명"
    ws1["B1"] = "값"
    _style_header_row(ws1, 1, 2)
    for i, (label, _key, hint) in enumerate(BASIC_FIELDS, start=2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=hint if hint else None)
        _style_label_cell(ws1.cell(row=i, column=1))
        _style_value_cell(ws1.cell(row=i, column=2))

    # ── 시트 2: 정량데이터 ──
    ws2 = wb.create_sheet("정량데이터")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 18
    ws2["A1"] = "필드명"
    ws2["B1"] = "값"
    _style_header_row(ws2, 1, 2)
    for i, (label, _key) in enumerate(QUANTITATIVE_FIELDS, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=0)
        _style_label_cell(ws2.cell(row=i, column=1))
        _style_value_cell(ws2.cell(row=i, column=2))

    # ── 시트 3: 프로그램 ──
    ws3 = wb.create_sheet("프로그램")
    headers = ["구분", "제목", "일자(YYYY-MM-DD)", "참여 인원", "비고"]
    widths = [16, 36, 18, 14, 28]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws3.cell(row=1, column=i, value=h)
        ws3.column_dimensions[get_column_letter(i)].width = w
    _style_header_row(ws3, 1, len(headers))
    # 예시 행 1개
    sample_prog = ["강연", "(예시) 색의 정치학: 노란색을 둘러싼 담론", "2025-06-20", 95, "큐레이터 토크"]
    for i, val in enumerate(sample_prog, start=1):
        c = ws3.cell(row=2, column=i, value=val)
        _style_value_cell(c)

    # ── 시트 4: 인쇄물 ──
    ws4 = wb.create_sheet("인쇄물")
    headers = ["종류", "수량", "비고"]
    widths = [18, 14, 50]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws4.cell(row=1, column=i, value=h)
        ws4.column_dimensions[get_column_letter(i)].width = w
    _style_header_row(ws4, 1, len(headers))
    sample_mat = ["리플렛", 8000, "4단 접지"]
    for i, val in enumerate(sample_mat, start=1):
        c = ws4.cell(row=2, column=i, value=val)
        _style_value_cell(c)

    # ── 시트 5: 언론보도 ──
    ws5 = wb.create_sheet("언론보도")
    headers = ["구분(인쇄/온라인)", "매체명", "일자(YYYY-MM-DD)", "제목", "비고/URL"]
    widths = [16, 18, 18, 50, 40]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws5.cell(row=1, column=i, value=h)
        ws5.column_dimensions[get_column_letter(i)].width = w
    _style_header_row(ws5, 1, len(headers))
    sample_press = ["인쇄", "조선일보", "2025-06-12", "(예시) 노란빛으로 물든 미술관", "문화면"]
    for i, val in enumerate(sample_press, start=1):
        c = ws5.cell(row=2, column=i, value=val)
        _style_value_cell(c)

    # ── 시트 6: 관객후기 ──
    ws6 = wb.create_sheet("관객후기")
    headers = ["분류(긍정/부정/기타)", "내용", "출처"]
    widths = [20, 60, 18]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws6.cell(row=1, column=i, value=h)
        ws6.column_dimensions[get_column_letter(i)].width = w
    _style_header_row(ws6, 1, len(headers))
    sample_rev = ["긍정", "(예시) 전시장 전체가 노란빛으로 물든 경험이 압도적이었습니다.", "방명록"]
    for i, val in enumerate(sample_rev, start=1):
        c = ws6.cell(row=2, column=i, value=val)
        _style_value_cell(c)

    # 안내 시트 (맨 앞)
    ws_help = wb.create_sheet("README", 0)
    ws_help.column_dimensions["A"].width = 90
    ws_help["A1"] = "일민미술관 전시 워크스페이스 — 일괄 임포트 템플릿"
    ws_help["A1"].font = Font(bold=True, size=14, color="255C4A")
    instructions = [
        "",
        "이 파일은 한 건의 전시 데이터를 작성하기 위한 Excel 템플릿입니다.",
        "각 시트의 필드에 값을 입력한 뒤, 워크스페이스 화면에서 'Excel 가져오기' 버튼으로 업로드하세요.",
        "",
        "▸ 기본정보 시트 — 제목·기간·작가·기획진·서술 항목 (B열에 값을 채우세요)",
        "▸ 정량데이터 시트 — 예산·관객·작품·프로그램·인력 등 숫자 (B열의 0을 실제 값으로 수정)",
        "▸ 프로그램 시트 — 한 행 = 한 프로그램. 필요한 만큼 행을 추가하세요",
        "▸ 인쇄물 시트 — 종류별로 한 행씩",
        "▸ 언론보도 시트 — A열 '인쇄' 또는 '온라인' 구분 명시",
        "▸ 관객후기 시트 — A열 '긍정' / '부정' / '기타' 분류",
        "",
        "주의:",
        "  - 날짜는 YYYY-MM-DD 형식을 권장합니다 (Excel 날짜 형식도 자동 인식)",
        "  - 빈 값은 입력하지 않아도 됩니다 (시스템이 0 또는 빈 문자열로 처리)",
        "  - 시트 이름은 변경하지 마세요",
        "",
        "업로드 후 폼에 미리 채워진 상태로 진입하므로 검수·수정이 가능합니다.",
        "최종 저장은 '워크스페이스에 저장' 버튼으로 직접 실행해야 합니다.",
    ]
    for i, line in enumerate(instructions, start=2):
        cell = ws_help.cell(row=i, column=1, value=line)
        cell.font = Font(size=10, color="20231F" if line else "FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 바이트로 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# 파싱
# ──────────────────────────────────────────────

def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_num(v):
    """셀 값 → 숫자(0 fallback)."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v) if isinstance(v, float) and v.is_integer() else v
    try:
        s = str(v).replace(",", "").replace("원", "").strip()
        if not s:
            return 0
        f = float(s)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return 0


def _to_date_str(v) -> Optional[str]:
    """다양한 날짜 표현 → 'YYYY-MM-DD' or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, _date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # 점/슬래시 구분자 정규화
    s = s.replace(".", "-").replace("/", "-")
    # YYYY-MM-DD 또는 YYYY-M-D
    parts = s.split("-")
    if len(parts) == 3:
        try:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            return f"{y:04d}-{m:02d}-{d:02d}"
        except ValueError:
            return None
    return None


def parse_template_xlsx(file) -> dict:
    """업로드된 Excel 템플릿을 v5 레코드 data 형식으로 변환.

    Returns: {
        "data": dict,           # session_state에 주입할 평면 dict
        "type": int | None,     # 전시 유형 (top-level)
        "warnings": list[str],  # 파싱 중 경고 메시지
    }
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    data = {}
    warnings = []
    exhibition_type = None

    # ── 시트 1: 기본정보 ──
    if "기본정보" in wb.sheetnames:
        ws = wb["기본정보"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            label, value = row[0], row[1]
            if not label:
                continue
            # 라벨 → key 매핑
            for lab, key, _hint in BASIC_FIELDS:
                if str(label).strip() == lab:
                    if key == "__exhibition_type":
                        if value not in (None, "", 0):
                            try:
                                exhibition_type = int(value)
                            except (ValueError, TypeError):
                                warnings.append(f"전시 유형 값 인식 실패: '{value}' (1/2/3 중 선택)")
                    elif key in ("period_start", "period_end"):
                        data[key] = _to_date_str(value)
                    else:
                        data[key] = _to_str(value)
                    break
    else:
        warnings.append("'기본정보' 시트가 없습니다.")

    # ── 시트 2: 정량데이터 ──
    if "정량데이터" in wb.sheetnames:
        ws = wb["정량데이터"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            label, value = row[0], row[1]
            if not label:
                continue
            for lab, key in QUANTITATIVE_FIELDS:
                if str(label).strip() == lab:
                    data[key] = _to_num(value)
                    break

    # ── 시트 3: 프로그램 ──
    if "프로그램" in wb.sheetnames:
        ws = wb["프로그램"]
        programs = []
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            cat, title, date, participants, note = row
            if not (cat or title):
                continue
            # 예시 행 자동 스킵 ("(예시)" 포함)
            if title and "(예시)" in str(title):
                continue
            programs.append({
                "category": _to_str(cat),
                "title": _to_str(title),
                "date": _to_date_str(date),
                "participants": _to_str(participants),
                "note": _to_str(note),
            })
        if programs:
            data["related_programs"] = programs

    # ── 시트 4: 인쇄물 ──
    if "인쇄물" in wb.sheetnames:
        ws = wb["인쇄물"]
        materials = []
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            mtype, qty, note = row
            if not (mtype or qty):
                continue
            materials.append({
                "type": _to_str(mtype),
                "quantity": _to_str(qty),
                "note": _to_str(note),
            })
        if materials:
            data["printed_materials"] = materials

    # ── 시트 5: 언론보도 ──
    if "언론보도" in wb.sheetnames:
        ws = wb["언론보도"]
        press_print = []
        press_online = []
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            ptype, outlet, date, title, note_or_url = row
            if not (outlet or title):
                continue
            if title and "(예시)" in str(title):
                continue
            ptype_s = (_to_str(ptype) or "").lower()
            entry = {
                "outlet": _to_str(outlet),
                "date": _to_date_str(date),
                "title": _to_str(title),
            }
            if "온라인" in ptype_s or "online" in ptype_s:
                entry["url"] = _to_str(note_or_url)
                press_online.append(entry)
            else:
                entry["note"] = _to_str(note_or_url)
                press_print.append(entry)
        if press_print:
            data["press_print"] = press_print
        if press_online:
            data["press_online"] = press_online

    # ── 시트 6: 관객후기 ──
    if "관객후기" in wb.sheetnames:
        ws = wb["관객후기"]
        reviews = []
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            cat, content, source = row
            if not content:
                continue
            if "(예시)" in str(content):
                continue
            reviews.append({
                "category": _to_str(cat) or "긍정",
                "content": _to_str(content),
                "source": _to_str(source),
            })
        if reviews:
            data["visitor_reviews"] = reviews

    # 뉴스레터 오픈율: 소수면 percent로 변환
    if data.get("newsletter_open_rate"):
        v = data["newsletter_open_rate"]
        if 0 < v < 1:
            data["newsletter_open_rate"] = round(v * 100, 1)

    return {
        "data": data,
        "type": exhibition_type,
        "warnings": warnings,
    }
