"""탭 A: 정량 데이터 — 분석의 대상이 되는 모든 숫자"""

import streamlit as st
import pandas as pd
import os


def render(tab):
    with tab:
        st.markdown('<div class="section-header">📊 정량 데이터</div>', unsafe_allow_html=True)
        st.caption("숫자를 입력하면 다음 탭에서 과거 전시와의 비교 분석이 자동 생성됩니다.")

        # ════════════════════════════════════════
        # 1. 예산
        # ════════════════════════════════════════
        st.subheader("💰 예산 및 수입")

        col1, col2 = st.columns(2)
        with col1:
            st.number_input("전시 사용 예산 (원)", min_value=0, step=1_000_000,
                            key="budget_exhibition", format="%d")
        with col2:
            st.number_input("부대 사용 예산 (원)", min_value=0, step=100_000,
                            key="budget_supplementary", format="%d")

        # 총 사용 예산 자동 합산
        total_budget = st.session_state.budget_exhibition + st.session_state.budget_supplementary
        st.session_state.total_budget = total_budget
        if total_budget > 0:
            st.metric("총 사용 예산", f"{total_budget:,}원")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.number_input("예산 계획액 (원)", min_value=0, step=1_000_000,
                            key="budget_planned", format="%d")
        with col2:
            st.number_input("입장 수입 (원)", min_value=0, step=100_000,
                            key="ticket_revenue", format="%d")
        with col3:
            st.number_input("기타 수입 (원)", min_value=0, step=100_000,
                            key="other_revenue", format="%d")

        # 총수입 자동 합산
        total_revenue = st.session_state.ticket_revenue + st.session_state.other_revenue
        st.session_state.total_revenue = total_revenue
        if total_revenue > 0:
            st.metric("총수입", f"{total_revenue:,}원")

        # ── 전시 데이터 엑셀 업로드 (예산 + 언론보도) ──
        st.markdown("**전시 데이터 엑셀 업로드**")
        st.caption("표준 엑셀 템플릿에 예산 집행 내역과 언론보도 리스트를 작성한 뒤 업로드하면, 보고서에 자동 반영됩니다.")

        col_dl, col_ul = st.columns(2)
        with col_dl:
            if st.button("📥 데이터 템플릿 다운로드", key="dl_data_tpl"):
                _create_data_template()
        with col_ul:
            data_file = st.file_uploader("전시 데이터 엑셀 업로드", type=["xlsx", "xls"],
                                          key="data_excel_upload")
            if data_file:
                _process_data_excel(data_file)

        # 업로드된 예산 데이터 미리보기
        if st.session_state.get("budget_summary") and any(
                x.get("category") for x in st.session_state.budget_summary):
            with st.expander("📋 업로드된 예산 집행 내역", expanded=False):
                summary_df = pd.DataFrame(st.session_state.budget_summary)
                summary_df = summary_df[summary_df["category"].astype(bool)]
                st.dataframe(summary_df, use_container_width=True, hide_index=True)

        if st.session_state.get("budget_details") and any(
                x.get("subcategory") or x.get("detail") for x in st.session_state.budget_details):
            with st.expander("📋 업로드된 예산 세부 내역", expanded=False):
                details_df = pd.DataFrame(st.session_state.budget_details)
                details_df = details_df[details_df["subcategory"].astype(bool) | details_df["detail"].astype(bool)]
                st.dataframe(details_df, use_container_width=True, hide_index=True)

        # 업로드된 언론보도 미리보기
        _press_print_from_excel = [p for p in st.session_state.get("press_print", []) if p.get("outlet")]
        _press_online_from_excel = [p for p in st.session_state.get("press_online", []) if p.get("outlet")]
        if _press_print_from_excel or _press_online_from_excel:
            with st.expander(f"📰 업로드된 언론보도 ({len(_press_print_from_excel) + len(_press_online_from_excel)}건)", expanded=False):
                if _press_print_from_excel:
                    st.markdown("**일간지/월간지**")
                    pp_df = pd.DataFrame(_press_print_from_excel)
                    st.dataframe(pp_df, use_container_width=True, hide_index=True)
                if _press_online_from_excel:
                    st.markdown("**온라인 매체**")
                    po_df = pd.DataFrame(_press_online_from_excel)
                    st.dataframe(po_df, use_container_width=True, hide_index=True)

        # 자동 계산 표시
        if total_budget > 0:
            metrics = st.columns(3)
            with metrics[0]:
                if st.session_state.budget_exhibition and st.session_state.budget_supplementary:
                    ratio = st.session_state.budget_exhibition / total_budget * 100
                    st.metric("전시비 비율", f"{ratio:.1f}%")
            with metrics[1]:
                if total_revenue:
                    recovery = total_revenue / total_budget * 100
                    st.metric("예산 회수율", f"{recovery:.1f}%")
            with metrics[2]:
                if st.session_state.budget_planned:
                    exec_rate = total_budget / st.session_state.budget_planned * 100
                    st.metric("집행률", f"{exec_rate:.1f}%")

        st.divider()

        # ════════════════════════════════════════
        # 2. 관객
        # ════════════════════════════════════════
        st.subheader("👥 관객")

        col1, col2 = st.columns(2)
        with col1:
            st.number_input("총 관객수", min_value=0, step=100,
                            key="total_visitors", format="%d")
        with col2:
            # 일평균 자동 계산
            days = None
            if st.session_state.period_start and st.session_state.period_end:
                days = (st.session_state.period_end - st.session_state.period_start).days + 1
            if st.session_state.total_visitors and days and days > 0:
                daily_avg = st.session_state.total_visitors // days
                st.metric("일평균 관객수 (자동)", f"{daily_avg:,}명")
            else:
                st.caption("일평균 관객수: 전시 기간 입력 시 자동 계산")

        st.markdown("**입장권별 관객 구성**")
        cols = st.columns(6)
        with cols[0]:
            st.number_input("일반", min_value=0, key="visitor_general", format="%d")
        with cols[1]:
            st.number_input("학생", min_value=0, key="visitor_student", format="%d")
        with cols[2]:
            st.number_input("초대권", min_value=0, key="visitor_invitation", format="%d")
        with cols[3]:
            st.number_input("예술인패스", min_value=0, key="visitor_artpass", format="%d")
        with cols[4]:
            st.number_input("디스커버서울패스", min_value=0, key="visitor_discover", format="%d")
        with cols[5]:
            st.number_input("기타 할인", min_value=0, key="visitor_discount", format="%d")

        # 합계 자동 검증
        ticket_sum = (st.session_state.visitor_general + st.session_state.visitor_student +
                      st.session_state.visitor_invitation + st.session_state.visitor_artpass +
                      st.session_state.visitor_discover + st.session_state.visitor_discount)
        if ticket_sum > 0 and st.session_state.total_visitors > 0:
            if ticket_sum != st.session_state.total_visitors:
                st.warning(f"⚠️ 입장권별 합계({ticket_sum:,}명)와 총 관객수({st.session_state.total_visitors:,}명)가 다릅니다.")
            else:
                st.success(f"✅ 입장권별 합계 일치: {ticket_sum:,}명")

        col1, col2 = st.columns(2)
        with col1:
            st.number_input("단체 관객수", min_value=0, key="visitor_group", format="%d")
        with col2:
            st.number_input("오프닝 참석 인원", min_value=0, key="opening_attendance", format="%d")

        # ── 주차별 관객 수 ──
        st.markdown("**주차별 관객 수**")
        st.caption("보고서에 주차별 추이 차트가 삽입됩니다. 빈 칸은 무시됩니다.")
        week_cols = st.columns(6)
        weekly = st.session_state.get("weekly_visitors", {})
        new_weekly = {}
        for i in range(6):
            with week_cols[i]:
                label = f"{i+1}주"
                val = weekly.get(label, 0)
                entered = st.number_input(label, min_value=0, value=val,
                                          key=f"weekly_{i}", format="%d")
                if entered > 0:
                    new_weekly[label] = entered
        st.session_state.weekly_visitors = new_weekly

        st.divider()

        # ════════════════════════════════════════
        # 3. 출품 작품
        # ════════════════════════════════════════
        st.subheader("🎨 출품 작품")

        cols = st.columns(6)
        with cols[0]:
            st.number_input("회화", min_value=0, key="artwork_painting", format="%d")
        with cols[1]:
            st.number_input("조각", min_value=0, key="artwork_sculpture", format="%d")
        with cols[2]:
            st.number_input("사진", min_value=0, key="artwork_photo", format="%d")
        with cols[3]:
            st.number_input("설치", min_value=0, key="artwork_installation", format="%d")
        with cols[4]:
            st.number_input("미디어", min_value=0, key="artwork_media", format="%d")
        with cols[5]:
            st.number_input("기타", min_value=0, key="artwork_other", format="%d")

        # 출품 작품 수 자동 합산
        artwork_total = (st.session_state.artwork_painting + st.session_state.artwork_sculpture +
                         st.session_state.artwork_photo + st.session_state.artwork_installation +
                         st.session_state.artwork_media + st.session_state.artwork_other)
        st.session_state.artwork_total = artwork_total
        if artwork_total > 0:
            st.metric("출품 작품 수 (총)", f"{artwork_total}점")

        st.divider()

        # ════════════════════════════════════════
        # 4. 프로그램
        # ════════════════════════════════════════
        st.subheader("🎯 프로그램 & 도슨트")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.number_input("프로그램 총 수", min_value=0, key="program_count", format="%d")
        with col2:
            st.number_input("프로그램 총 회차", min_value=0, key="program_sessions", format="%d")
        with col3:
            st.number_input("프로그램 참여 인원", min_value=0, key="program_participants", format="%d")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.number_input("도슨트 참여 인원 (총)", min_value=0, key="docent_total", format="%d")
        with col2:
            st.number_input("정기 도슨트", min_value=0, key="docent_regular", format="%d")
        with col3:
            st.number_input("특별 도슨트", min_value=0, key="docent_special", format="%d")

        st.divider()

        # ════════════════════════════════════════
        # 5. 운영 인력
        # ════════════════════════════════════════
        st.subheader("👷 운영 인력")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.number_input("운영 인력 총원", min_value=0, key="staff_total", format="%d")
        with col2:
            st.number_input("유급 스태프", min_value=0, key="staff_paid", format="%d")
        with col3:
            st.number_input("봉사자", min_value=0, key="staff_volunteer", format="%d")

        st.divider()

        # ════════════════════════════════════════
        # 6. 홍보 지표
        # ════════════════════════════════════════
        st.subheader("📢 홍보 지표")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.number_input("언론 보도 건수", min_value=0, key="press_count", format="%d",
                            help="일간지+온라인 합계. '기본 정보' 탭의 보도 리스트와 연동됩니다.")
        with col2:
            st.number_input("웹 초청장 발송 수", min_value=0, key="web_invitation_count", format="%d")
        with col3:
            st.number_input("뉴스레터 오픈율 (%)", min_value=0.0, max_value=100.0,
                            step=0.1, key="newsletter_open_rate", format="%.1f")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.number_input("SNS 게시 건수", min_value=0, key="sns_posts", format="%d")
        with col2:
            st.number_input("SNS 피드백 합계", min_value=0, key="sns_feedback", format="%d")
        with col3:
            st.number_input("멤버십 회원수", min_value=0, key="membership_count", format="%d")

        # ── SNS 상세 통계 ──
        st.markdown("**SNS 상세 통계**")
        st.caption("인스타그램 기준 정량 지표입니다. 보고서 홍보 섹션에 반영됩니다.")
        sns_cols = st.columns(4)
        with sns_cols[0]:
            st.number_input("팔로워 수", min_value=0, key="sns_followers", format="%d")
        with sns_cols[1]:
            st.number_input("팔로워 증가", min_value=0, key="sns_followers_gained", format="%d",
                            help="전시 기간 중 순증가 수")
        with sns_cols[2]:
            st.number_input("평균 좋아요", min_value=0, key="sns_avg_likes", format="%d")
        with sns_cols[3]:
            st.number_input("최고 좋아요", min_value=0, key="sns_best_likes", format="%d")
        st.text_input("최고 반응 게시물 내용", key="sns_best_post",
                      placeholder="예: 한강주조 겨울 에디션 게시물")

        # ── 보도 건수 자동 동기화 제안 ──
        print_count = len([p for p in st.session_state.press_print if p.get("outlet")])
        online_count = len([p for p in st.session_state.press_online if p.get("outlet")])
        list_total = print_count + online_count
        if list_total > 0 and st.session_state.press_count == 0:
            st.info(f"💡 '기본 정보' 탭에 보도 {list_total}건이 입력되어 있습니다. 언론 보도 건수를 {list_total}으로 설정하시겠습니까?")
            if st.button("자동 입력", key="sync_press"):
                st.session_state.press_count = list_total
                st.rerun()


# ──────────────────────────────────────────────
# 예산 엑셀 템플릿 생성 및 파싱
# ──────────────────────────────────────────────

def _create_data_template():
    """전시 데이터 표준 엑셀 템플릿 생성 (예산 + 언론보도)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

        def _style_header(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        def _fill_rows(ws, rows, start_row=2, money_cols=None):
            for r, row_data in enumerate(rows, start_row):
                for c, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                    if money_cols and c in money_cols:
                        cell.number_format = '#,##0'

        # ── Sheet 1: 예산 요약 ──
        ws1 = wb.active
        ws1.title = "예산 요약"
        _style_header(ws1, ["사업", "계획 예산(원)", "집행 예산(원)", "비고"])
        _fill_rows(ws1, [
            ["전시 제작", "", "", ""],
            ["전시 부대", "", "", ""],
            ["합계", "", "", ""],
        ], money_cols={2, 3})
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 18
        ws1.column_dimensions["C"].width = 18
        ws1.column_dimensions["D"].width = 25

        # ── Sheet 2: 세부 내역 ──
        ws2 = wb.create_sheet("세부 내역")
        _style_header(ws2, ["사업 구분", "항목", "세부 내용", "금액(원)", "비고"])
        _fill_rows(ws2, [
            ["전시 제작", "공간 조성", "전시실 가벽, 조명, 페인트", "", ""],
            ["전시 제작", "작품 운송", "국내 운송비", "", ""],
            ["전시 제작", "인쇄물", "포스터, 리플릿, 티켓", "", ""],
            ["전시 부대", "연계 프로그램", "아티스트 토크, 워크숍", "", ""],
            ["전시 부대", "홍보비", "온라인 광고, SNS", "", ""],
        ], money_cols={4})
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 30
        ws2.column_dimensions["D"].width = 18
        ws2.column_dimensions["E"].width = 20

        # ── Sheet 3: 언론보도 일간지 ──
        ws3 = wb.create_sheet("언론보도 일간지")
        _style_header(ws3, ["매체명", "보도 일자", "기사 제목", "비고"])
        _fill_rows(ws3, [
            ["한겨레", "2025-09-15", "일민미술관, 가을 기획전 개막", ""],
            ["조선일보", "2025-10-01", "현대미술의 새 지평", "문화면"],
        ])
        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 15
        ws3.column_dimensions["C"].width = 45
        ws3.column_dimensions["D"].width = 20

        # ── Sheet 4: 언론보도 온라인 ──
        ws4 = wb.create_sheet("언론보도 온라인")
        _style_header(ws4, ["매체명", "보도 일자", "기사 제목", "URL"])
        _fill_rows(ws4, [
            ["아트인사이트", "2025-09-16", "일민미술관 전시 리뷰", "https://example.com/article1"],
            ["네오룩", "2025-09-20", "이번 주 주목할 전시", "https://example.com/article2"],
        ])
        ws4.column_dimensions["A"].width = 18
        ws4.column_dimensions["B"].width = 15
        ws4.column_dimensions["C"].width = 45
        ws4.column_dimensions["D"].width = 40

        # 바이트로 변환하여 다운로드
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            "📥 템플릿 다운로드",
            buffer.getvalue(),
            file_name="전시_데이터_템플릿.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ImportError:
        st.error("openpyxl 패키지가 필요합니다. `pip install openpyxl`을 실행하세요.")


def _process_data_excel(uploaded_file):
    """업로드된 전시 데이터 엑셀을 파싱하여 세션에 저장 (예산 + 언론보도)"""
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names

        def _clean(val):
            """NaN 및 빈 문자열 정리"""
            s = str(val).strip() if val is not None else ""
            return "" if s == "nan" or s == "None" else s

        def _money_fmt(val):
            """숫자를 천 단위 구분 문자열로"""
            if isinstance(val, (int, float)) and not pd.isna(val):
                return f"{int(val):,}"
            s = _clean(val)
            return s

        # ── Sheet 1: 예산 요약 ──
        if "예산 요약" in sheet_names or len(sheet_names) >= 1:
            sn = "예산 요약" if "예산 요약" in sheet_names else sheet_names[0]
            df = pd.read_excel(xls, sheet_name=sn)
            col_map = {}
            for col in df.columns:
                cl = str(col).strip()
                if "사업" in cl or "구분" in cl:
                    col_map[col] = "category"
                elif "계획" in cl:
                    col_map[col] = "planned"
                elif "집행" in cl or "실제" in cl:
                    col_map[col] = "actual"
                elif "비고" in cl or "메모" in cl:
                    col_map[col] = "note"
            if col_map:
                df = df.rename(columns=col_map)
                rows = []
                for _, row in df.iterrows():
                    cat = _clean(row.get("category", ""))
                    if not cat:
                        continue
                    rows.append({
                        "category": cat,
                        "planned": _money_fmt(row.get("planned", "")),
                        "actual": _money_fmt(row.get("actual", "")),
                        "note": _clean(row.get("note", "")),
                    })
                if rows:
                    st.session_state.budget_summary = rows
                    st.success(f"✅ 예산 요약 {len(rows)}건 로드")

        # ── Sheet 2: 세부 내역 ──
        if "세부 내역" in sheet_names or len(sheet_names) >= 2:
            sn = "세부 내역" if "세부 내역" in sheet_names else (sheet_names[1] if len(sheet_names) >= 2 else None)
            if sn:
                df = pd.read_excel(xls, sheet_name=sn)
                col_map = {}
                for col in df.columns:
                    cl = str(col).strip()
                    if "사업" in cl or "구분" in cl:
                        col_map[col] = "category"
                    elif "항목" in cl:
                        col_map[col] = "subcategory"
                    elif "세부" in cl or "내용" in cl:
                        col_map[col] = "detail"
                    elif "금액" in cl or "원" in cl:
                        col_map[col] = "amount"
                    elif "비고" in cl or "메모" in cl:
                        col_map[col] = "note"
                if col_map:
                    df = df.rename(columns=col_map)
                    rows = []
                    for _, row in df.iterrows():
                        subcat = _clean(row.get("subcategory", ""))
                        detail = _clean(row.get("detail", ""))
                        if not subcat and not detail:
                            continue
                        rows.append({
                            "category": _clean(row.get("category", "")),
                            "subcategory": subcat,
                            "detail": detail,
                            "amount": _money_fmt(row.get("amount", "")),
                            "note": _clean(row.get("note", "")),
                        })
                    if rows:
                        st.session_state.budget_details = rows
                        st.success(f"✅ 세부 내역 {len(rows)}건 로드")

        # ── Sheet 3: 언론보도 일간지 ──
        _process_press_sheet(xls, sheet_names, "언론보도 일간지", "press_print",
                             has_url=False)

        # ── Sheet 4: 언론보도 온라인 ──
        _process_press_sheet(xls, sheet_names, "언론보도 온라인", "press_online",
                             has_url=True)

    except Exception as e:
        st.error(f"엑셀 파싱 오류: {e}")


def _process_press_sheet(xls, sheet_names, target_sheet, session_key, has_url=False):
    """언론보도 시트 파싱 공통 함수"""
    if target_sheet not in sheet_names:
        return

    df = pd.read_excel(xls, sheet_name=target_sheet)
    col_map = {}
    for col in df.columns:
        cl = str(col).strip()
        if "매체" in cl:
            col_map[col] = "outlet"
        elif "일자" in cl or "날짜" in cl or "보도" in cl:
            col_map[col] = "date"
        elif "제목" in cl or "기사" in cl:
            col_map[col] = "title"
        elif "url" in cl.lower() or "링크" in cl:
            col_map[col] = "url"
        elif "비고" in cl or "메모" in cl:
            col_map[col] = "note"

    if not col_map:
        return

    df = df.rename(columns=col_map)
    rows = []
    for _, row in df.iterrows():
        outlet = str(row.get("outlet", "")).strip()
        if not outlet or outlet == "nan":
            continue

        # 날짜 처리
        date_val = row.get("date")
        if pd.notna(date_val):
            try:
                if hasattr(date_val, 'date'):
                    date_val = date_val.date()
                elif isinstance(date_val, str):
                    from datetime import date as dt_date
                    date_val = dt_date.fromisoformat(date_val)
            except (ValueError, AttributeError):
                date_val = None
        else:
            date_val = None

        title = str(row.get("title", "")).strip()
        title = "" if title == "nan" else title

        entry = {"outlet": outlet, "date": date_val, "title": title}
        if has_url:
            url = str(row.get("url", "")).strip()
            entry["url"] = "" if url == "nan" else url
        else:
            note = str(row.get("note", "")).strip()
            entry["note"] = "" if note == "nan" else note

        rows.append(entry)

    if rows:
        st.session_state[session_key] = rows
        label = "일간지/월간지" if not has_url else "온라인 매체"
        st.success(f"✅ 언론보도 {label} {len(rows)}건 로드")
