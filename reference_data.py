"""
레퍼런스 데이터 관리 모듈
- Excel 파일에서 과거 전시 데이터 로드
- 통계 계산 (평균, 중앙값, 백분위 등)
- 유사 전시 검색
- 새 전시 데이터 추가/저장
"""

import os
import numpy as np
import pandas as pd
from dataclasses import dataclass


# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────

# Excel 구조: Row 1 = 카테고리 헤더 (병합), Row 2 = 컬럼명, Row 3+ = 데이터
HEADER_ROW = 1  # 0-indexed for pandas (실제 Excel Row 2)
DATA_START_ROW = 2  # 0-indexed (실제 Excel Row 3)

# 숫자로 파싱해야 하는 컬럼들
NUMERIC_COLUMNS = [
    "전시 일수",
    "참여 작가 수_총(팀)", "참여 작가 수_국내", "참여 작가 수_해외", "참여 작가 수_소장 작가",
    "총 사용 예산", "전시 사용 예산", "부대 사용 예산", "예산 계획액", "예산 집행률",
    "총수입", "입장 수입",
    "총 관객수", "일평균 관객수", "유료 관객수", "무료/초대 관객수",
    "학생 관객수(만 24세 이하)", "단체 관객수", "디스커버서울패스 관객수", "예술인패스 관객수",
    "운영 인력_총", "스태프 수", "봉사자 수", "지원단 수",
    "프로그램 총 수", "프로그램 총 회차", "프로그램 참여 인원",
    "도슨트 참여 인원", "정기 도슨트 참여 인원", "특별 도슨트 참여 인원", "오프닝 참석 인원",
    "출품 작품 수_총", "출품 작품 수_신작", "출품 작품 수_구작",
    "출품 작품 수_회화", "출품 작품 수_조각", "출품 작품 수_사진",
    "출품 작품 수_설치", "출품 작품 수_미디어", "출품 작품 수_기타",
    "언론 보도 건수", "웹 초청장 발송 수", "뉴스레터 오픈율",
    "SNS 게시 건수", "SNS 피드백 합계",
    "멤버십 회원수", "멤버십 증가율",
]

# 유사 전시 검색에 사용할 핵심 비교 필드 및 가중치
SIMILARITY_FIELDS = {
    "총 사용 예산": 0.35,
    "전시 일수": 0.25,
    "총 관객수": 0.25,
    "참여 작가 수_총(팀)": 0.15,
}


# ──────────────────────────────────────────────
# Excel 로드
# ──────────────────────────────────────────────

def load_reference(xlsx_path: str) -> pd.DataFrame:
    """
    레퍼런스 Excel 파일을 DataFrame으로 로드.

    Excel 구조:
        Row 1: 카테고리 헤더 (병합 셀) — 건너뜀
        Row 2: 실제 컬럼명
        Row 3+: 전시 데이터

    Returns:
        pd.DataFrame: 전시별 한 행, 컬럼명은 Row 2 기준
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"레퍼런스 파일을 찾을 수 없습니다: {xlsx_path}")

    # Row 1(0-indexed) = 컬럼명, skiprows=[0]으로 카테고리 헤더 건너뜀
    df = pd.read_excel(
        xlsx_path,
        sheet_name=0,
        header=1,      # Row 2 (0-indexed=1) 를 컬럼명으로
        engine="openpyxl",
    )

    # 빈 행 제거 (전시 제목이 없는 행)
    df = df.dropna(subset=["전시 제목"], how="all")
    df = df[df["전시 제목"].astype(str).str.strip() != ""]

    # '-' 문자를 NaN으로 변환
    df = df.replace("-", np.nan)
    df = df.replace("—", np.nan)
    df = df.replace("", np.nan)

    # 숫자 컬럼 변환
    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # 전시 유형 컬럼 숫자 변환
    if "전시 유형" in df.columns:
        df["전시 유형"] = pd.to_numeric(df["전시 유형"], errors="coerce")

    # No. 컬럼 정리
    if "No." in df.columns:
        df["No."] = range(1, len(df) + 1)

    df = df.reset_index(drop=True)
    return df


# ──────────────────────────────────────────────
# 통계 계산
# ──────────────────────────────────────────────

@dataclass
class FieldStats:
    """특정 필드의 통계 정보"""
    field_name: str
    count: int          # 유효 값 개수
    mean: float
    median: float
    min_val: float
    max_val: float
    std: float
    q25: float          # 25th percentile
    q75: float          # 75th percentile
    values: list        # 전체 유효 값 리스트 (순위 계산용)
    titles: list        # 대응하는 전시 제목 리스트


def compute_stats(df: pd.DataFrame, column: str) -> FieldStats | None:
    """
    특정 필드의 통계를 계산합니다.

    Args:
        df: 레퍼런스 DataFrame
        column: 필드명

    Returns:
        FieldStats 또는 None (필드가 없거나 데이터 부족 시)
    """
    if column not in df.columns:
        return None

    series = df[column].dropna()
    if len(series) < 2:
        return None

    # 유효 값과 대응하는 전시 제목
    valid_mask = df[column].notna()
    values = df.loc[valid_mask, column].tolist()
    titles = df.loc[valid_mask, "전시 제목"].tolist()

    return FieldStats(
        field_name=column,
        count=len(series),
        mean=float(series.mean()),
        median=float(series.median()),
        min_val=float(series.min()),
        max_val=float(series.max()),
        std=float(series.std()),
        q25=float(series.quantile(0.25)),
        q75=float(series.quantile(0.75)),
        values=values,
        titles=titles,
    )


def compute_percentile(stats: FieldStats, value: float) -> int:
    """
    주어진 값의 백분위를 계산합니다 (0-100).
    """
    if stats is None or stats.count == 0:
        return 50
    sorted_vals = sorted(stats.values)
    below = sum(1 for v in sorted_vals if v < value)
    equal = sum(1 for v in sorted_vals if v == value)
    percentile = (below + equal * 0.5) / len(sorted_vals) * 100
    return int(round(percentile))


def compute_rank(stats: FieldStats, value: float, ascending: bool = False) -> int:
    """
    주어진 값의 순위를 계산합니다.

    Args:
        ascending: True이면 낮을수록 좋은 순위 (예: 관객당 비용)
    """
    if stats is None or stats.count == 0:
        return 0
    if ascending:
        sorted_vals = sorted(stats.values)
    else:
        sorted_vals = sorted(stats.values, reverse=True)
    # 같은 값이면 가장 좋은 순위
    for i, v in enumerate(sorted_vals):
        if abs(v - value) < 0.01 or v == value:
            return i + 1
    # 값이 리스트에 없으면 삽입 위치 기준
    for i, v in enumerate(sorted_vals):
        if (not ascending and value > v) or (ascending and value < v):
            return i + 1
    return len(sorted_vals) + 1


# ──────────────────────────────────────────────
# 유사 전시 검색
# ──────────────────────────────────────────────

def get_similar_exhibitions(
    df: pd.DataFrame,
    current: dict,
    top_n: int = 5,
) -> pd.DataFrame:
    """
    현재 전시와 유사한 과거 전시를 찾아 반환합니다.

    유사도 계산:
        각 비교 필드에 대해 정규화된 차이(0~1)를 계산하고
        가중 평균으로 종합 유사도 점수를 산출합니다.

    Args:
        df: 레퍼런스 DataFrame
        current: 현재 전시 데이터 dict (키: 필드명, 값: 숫자)
        top_n: 반환할 유사 전시 수

    Returns:
        유사도 순으로 정렬된 DataFrame (상위 top_n개)
    """
    if df.empty:
        return pd.DataFrame()

    scores = pd.Series(0.0, index=df.index)
    total_weight = 0.0

    for field, weight in SIMILARITY_FIELDS.items():
        if field not in current or current[field] is None:
            continue
        if field not in df.columns:
            continue

        current_val = float(current[field])
        if current_val == 0:
            continue

        col = df[field].copy()
        valid = col.notna() & (col != 0)

        if valid.sum() < 2:
            continue

        # 정규화된 차이 계산 (0 = 동일, 1 = 매우 다름)
        col_range = col[valid].max() - col[valid].min()
        if col_range == 0:
            continue

        diff = (col - current_val).abs() / col_range
        diff = diff.clip(upper=1.0)
        diff = diff.fillna(1.0)  # 데이터 없으면 최대 차이

        scores += diff * weight
        total_weight += weight

    if total_weight == 0:
        return df.head(top_n)

    scores = scores / total_weight  # 0~1 정규화
    df_result = df.copy()
    df_result["_similarity_score"] = 1 - scores  # 1에 가까울수록 유사
    df_result = df_result.sort_values("_similarity_score", ascending=False)

    return df_result.head(top_n)


# ──────────────────────────────────────────────
# 전시 유형 기반 필터링
# ──────────────────────────────────────────────

EXHIBITION_TYPE_COL = "전시 유형"
EXCLUDED_TYPE = 0  # 유형 0은 분석에서 제외 (특수 전시)


def exclude_type_zero(df: pd.DataFrame) -> pd.DataFrame:
    """유형 0(특수 전시)을 DataFrame에서 제외합니다."""
    if EXHIBITION_TYPE_COL not in df.columns:
        return df
    type_series = pd.to_numeric(df[EXHIBITION_TYPE_COL], errors="coerce")
    return df[type_series != EXCLUDED_TYPE]


def filter_by_type(df: pd.DataFrame, exhibition_type) -> pd.DataFrame:
    """
    동일 유형의 전시만 필터링합니다.
    유형 0은 항상 제외됩니다.

    Args:
        df: 레퍼런스 DataFrame
        exhibition_type: 유형 값 (1, 2, 3 등). None이면 유형 0 제외 후 전체 반환.

    Returns:
        필터링된 DataFrame. 같은 유형이 3개 미만이면 유형 0 제외 전체 반환.
    """
    # 유형 0은 항상 제외
    df = exclude_type_zero(df)

    if exhibition_type is None or EXHIBITION_TYPE_COL not in df.columns:
        return df

    # 유형 컬럼을 숫자로 변환
    type_series = pd.to_numeric(df[EXHIBITION_TYPE_COL], errors="coerce")
    try:
        target = float(exhibition_type)
    except (ValueError, TypeError):
        return df

    filtered = df[type_series == target]

    # 같은 유형이 3개 미만이면 의미 있는 비교가 어려우므로 전체 사용
    if len(filtered) < 3:
        return df

    return filtered


def get_type_label(exhibition_type) -> str:
    """유형 번호를 표시용 라벨로 변환합니다."""
    if exhibition_type is None:
        return "전체"
    try:
        t = int(float(exhibition_type))
    except (ValueError, TypeError):
        return "전체"
    return f"{t}유형"


def get_type_count(df: pd.DataFrame, exhibition_type) -> int:
    """특정 유형의 전시 수를 반환합니다."""
    if exhibition_type is None or EXHIBITION_TYPE_COL not in df.columns:
        return len(df)
    type_series = pd.to_numeric(df[EXHIBITION_TYPE_COL], errors="coerce")
    try:
        target = float(exhibition_type)
    except (ValueError, TypeError):
        return len(df)
    return int((type_series == target).sum())


# ──────────────────────────────────────────────
# 파생 지표 계산
# ──────────────────────────────────────────────

def compute_derived_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame에 파생 지표 컬럼을 추가합니다.

    추가되는 컬럼:
        - 관객당_비용: 총 사용 예산 ÷ 총 관객수
        - 수입_예산_비율: 총수입 ÷ 총 사용 예산
        - 유료_비율: 유료 관객수 ÷ 총 관객수
        - 프로그램_참여율: 프로그램 참여 인원 ÷ 총 관객수
        - 관객당_보도건수: 총 관객수 ÷ 언론 보도 건수
    """
    df = df.copy()

    # 관객당 비용 (원) — 필터링 없이 원본 비율 그대로 보존
    # (전시 유형별 비교로 이상값 문제를 해결)
    with np.errstate(divide="ignore", invalid="ignore"):
        df["관객당_비용"] = np.where(
            (df["총 관객수"].notna()) & (df["총 관객수"] > 0),
            df["총 사용 예산"] / df["총 관객수"],
            np.nan,
        )

        # 수입/예산 비율 — 필터링 없이 원본 비율 그대로 보존
        df["수입_예산_비율"] = np.where(
            (df["총 사용 예산"].notna()) & (df["총 사용 예산"] > 0),
            df["총수입"] / df["총 사용 예산"],
            np.nan,
        )

        # 유료 관객 비율
        df["유료_비율"] = np.where(
            (df["총 관객수"].notna()) & (df["총 관객수"] > 0),
            df["유료 관객수"] / df["총 관객수"],
            np.nan,
        )

        # 프로그램 참여율
        df["프로그램_참여율"] = np.where(
            (df["총 관객수"].notna()) & (df["총 관객수"] > 0),
            df["프로그램 참여 인원"] / df["총 관객수"],
            np.nan,
        )

        # 관객당 보도건수 (보도 1건당 관객)
        df["보도건당_관객"] = np.where(
            (df["언론 보도 건수"].notna()) & (df["언론 보도 건수"] > 0),
            df["총 관객수"] / df["언론 보도 건수"],
            np.nan,
        )

    return df


# ──────────────────────────────────────────────
# 현재 전시 데이터 → 비교용 dict 변환
# ──────────────────────────────────────────────

def app_data_to_reference_dict(app_data: dict) -> dict:
    """
    Streamlit app의 data dict를 레퍼런스 비교용 flat dict로 변환합니다.

    app.py의 collect_data()가 반환하는 중첩 구조를
    레퍼런스 Excel의 컬럼명에 맞춘 flat dict로 매핑합니다.
    """

    def parse_number(s):
        """문자열에서 숫자 추출"""
        if s is None:
            return None
        if isinstance(s, (int, float)):
            return float(s)
        s = str(s).replace(",", "").replace("명", "").replace("원", "").replace("%", "")
        s = s.replace("약 ", "").replace("개", "").replace("회", "").strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    overview = app_data.get("overview", {})
    budget = app_data.get("budget", {})
    revenue = app_data.get("revenue", {})
    visitor_comp = app_data.get("visitor_composition", {})

    # 관객 수 합산
    ticket = visitor_comp.get("ticket_type", {})
    vtype = visitor_comp.get("visitor_type", {})

    result = {
        "전시 제목": app_data.get("exhibition_title", ""),
        "전시 일수": parse_number(overview.get("exhibition_days"))
                     or (app_data.get("period_end", None) and None),  # 별도 계산 필요
        "총 사용 예산": parse_number(overview.get("total_budget")),
        "총수입": parse_number(revenue.get("total_revenue")),
        "입장 수입": parse_number(revenue.get("ticket_revenue")),
        "총 관객수": parse_number(revenue.get("total_visitors"))
                     or parse_number(overview.get("visitors")),
        "일평균 관객수": parse_number(revenue.get("daily_average")),
        "유료 관객수": ticket.get("일반", 0) + ticket.get("학생", 0),
        "무료/초대 관객수": ticket.get("초대권", 0),
        "학생 관객수(만 24세 이하)": ticket.get("학생", 0),
        "예술인패스 관객수": ticket.get("예술인패스", 0),
        "오프닝 참석 인원": vtype.get("오프닝 리셉션", 0),
        "단체 관객수": vtype.get("미술대학 단체", 0) + vtype.get("기타 단체", 0),
        "언론 보도 건수": (
            len(app_data.get("press_coverage", {}).get("print_media", []))
            + len(app_data.get("press_coverage", {}).get("online_media", []))
        ),
        "프로그램 총 수": len(app_data.get("related_programs", [])),
    }

    # 프로그램 참여 인원 합산
    total_participants = 0
    for prog in app_data.get("related_programs", []):
        p = parse_number(prog.get("participants"))
        if p:
            total_participants += p
    result["프로그램 참여 인원"] = total_participants or None

    return result


# ──────────────────────────────────────────────
# 레퍼런스 갱신
# ──────────────────────────────────────────────

def add_exhibition_to_reference(xlsx_path: str, new_data: dict):
    """
    새 전시 데이터를 레퍼런스 Excel에 추가합니다.

    Args:
        xlsx_path: Excel 파일 경로
        new_data: flat dict (키: 컬럼명, 값: 데이터)
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 컬럼명 → 컬럼 인덱스 매핑 (Row 2)
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=2, column=col_idx).value
        if header:
            col_map[str(header).strip()] = col_idx

    # 마지막 데이터 행 찾기
    last_row = 2
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value:
            last_row = row_idx
    new_row = last_row + 1

    # No. 설정
    if "No." in col_map:
        ws.cell(row=new_row, column=col_map["No."], value=new_row - 2)

    # 데이터 기입
    for field, value in new_data.items():
        if field in col_map and value is not None:
            ws.cell(row=new_row, column=col_map[field], value=value)

    wb.save(xlsx_path)


# ──────────────────────────────────────────────
# 포맷팅 유틸리티
# ──────────────────────────────────────────────

def format_number(value, unit: str = "") -> str:
    """숫자를 읽기 좋은 한국어 형식으로 변환"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "N/A"
    value = float(value)
    if abs(value) >= 1_0000_0000:
        return f"{value / 1_0000_0000:.1f}억{unit}"
    elif abs(value) >= 1_0000:
        return f"{value / 1_0000:.0f}만{unit}"
    elif abs(value) >= 1000:
        return f"{value:,.0f}{unit}"
    elif value == int(value):
        return f"{int(value)}{unit}"
    else:
        return f"{value:.1f}{unit}"


def format_percent(value) -> str:
    """비율을 퍼센트 문자열로 변환"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "N/A"
    return f"{float(value) * 100:.1f}%"
